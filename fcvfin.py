from __future__ import print_function
from liltilities import get_existing_sheets, show_files_as_choices, sheet_finder
from liltilities import Liltilities
from receipts import RentReceipts
from pprint import pprint
from datetime import datetime, timedelta, date
from config import RENT_SHEETS2022, CURRENT_YEAR_RS, READ_RANGE_HAP, READ_RANGE_PAY_PRE, R_RANGE_INTAKE_UNITS, my_scopes
from config import Config
from auth_work import oauth
from oauth2client.service_account import ServiceAccountCredentials
from google_api_calls_abstract import broad_get
import sys
import os
import pickle
import os.path
from time import sleep
import copy
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook, load_workbook
import gspread
from gspread_dataframe import get_as_dataframe, set_with_dataframe
from app.models import db, Unit, BasicBitchPayment

class UI(object):

    def __init__(self):
        pass

    def prompt(self, *args):
        for lines in args:
            pprint(lines)
        input1 = int(input('....>'))
        return input1
    # make this kwargs

class FileHandler(object):

    def __init__(self):
        self.spreadsheet_testbook = CURRENT_YEAR_RS #wb name: rent sheets 2020

class TemplateFormatSheet(object):

    def __init__(self):
        self.service = ''
        self.spreadsheet_id = ''
        self.read_range = ''

    def set_id(self, service, spreadsheet_id, read_range):
        self.service = service
        self.spreadsheet_id = spreadsheet_id
        self.read_range = read_range

    def batch_get(self, col_num): #
        service = self.service
        print(type(service))
        sheet = service.spreadsheets()
        read_range = self.read_range
        result = sheet.values().get(spreadsheetId=self.spreadsheet_id,
                                    range=self.read_range, majorDimension="ROWS").execute()

        values = result.get('values', [])
        col = []
        if not values:
            print('No data found.')
        else:
            for COLUMN in values:
                col.append(COLUMN[col_num])
                # print(COLUMN[col_num]) # this is just a view, not the complete object
        #print(col)
        return col

    # alt get implementation: you can pass in service and sheet (add majDim)
    def alt_batch_get(self, service, sheet_id, range):
        sheet = service.spreadsheets()
        result = sheet.values().get(spreadsheetId=sheet_id,
                                    range=range).execute()
        values = []
        values_ = result.get('values', [0])
        print('values:', values)
        try:
            for value in values_:
                print('value:', value)
                values.append(value[0])
        except TypeError as e:
            print(e, 'You probably need to put in the unit numbers into intake sheet.')
        return values

    def simple_batch_update(self, service, sheet_id, wrange, data, dim):
        pprint(f"Updating {sheet_id} with batch call to {wrange}...")
        sheet = service.spreadsheets()
        #range_ = wrange
        body_request = {
                        'value_input_option': 'RAW',
                        'data': [
                                {'range': wrange,
                                'majorDimension': dim,
                                 'values': [data]
                                 }
                        ],
                        }

        request = service.spreadsheets().values().batchUpdate(spreadsheetId=sheet_id, body=body_request)
        response = request.execute()
        print(response)

    def write_formula_column(self, data, write_range):
        pprint(f"Writing {data} to {write_range}  . . .")
        service = self.service
        sheet = service.spreadsheets()
        spreadsheet_id = self.spreadsheet_id
        range_ = write_range  # TODO: Update placeholder value.
        # How the input data should be interpreted.
        value_input_option = 'USER_ENTERED'
        value_range_body = {
                            "values": [data]
        }

        request = service.spreadsheets().values().update(spreadsheetId=spreadsheet_id, range=range_, valueInputOption=value_input_option, body=value_range_body)
        response = request.execute()

    def format_row(self, write_range, r_or_c, name_list): # and writing strings to ranges passed
        pprint("Formatting row . . .")
        service = self.service
        sheet = service.spreadsheets()
        spreadsheet_id = self.spreadsheet_id
        pprint(service)
        pprint(sheet)

        range_ = write_range  # TODO: Update placeholder value.

        # How the input data should be interpreted.
        value_input_option = 'USER_ENTERED'  #

        value_range_body = {"range": write_range,
                            "majorDimension": r_or_c,
                            "values": [name_list]
            # TODO: Add desired entries to the request body. All existing entries
            # will be replaced.
        }

        request = service.spreadsheets().values().update(spreadsheetId=spreadsheet_id, range=range_, valueInputOption=value_input_option, body=value_range_body)
        response = request.execute()
        print(response)
    def date_stamp(self, range):
        service = self.service
        sheet = service.spreadsheets()
        spreadsheet_id = self.spreadsheet_id
        d = ["Generated on:"]
        d.append(str(datetime.now()))

        range_ = range
        value_input_option = 'USER_ENTERED'
        value_range_body = {"range": range,
                        "values": [d]
                        }
        request = service.spreadsheets().values().update(spreadsheetId=spreadsheet_id, range=range_, valueInputOption=value_input_option, body=value_range_body)
        response = request.execute()
    # bolds 1 row, freezes num of rows
    def bold_freeze(self, sheet_id, num):
        service = self.service
        sheet = service.spreadsheets()
        spreadsheet_id = self.spreadsheet_id

        data = {"requests":
                [ {"repeatCell":
                    {"range": {
                        "sheetId": sheet_id,
                        "startRowIndex": 0,
                        "endRowIndex": 1},
                    "cell":  {
                        "userEnteredFormat": {
                            "textFormat": { "bold": True }}
                            },
                        "fields": "userEnteredFormat.textFormat.bold"}
                    },
                    {'updateSheetProperties': {
                        'properties': {
                            'sheetId': sheet_id,
                            'gridProperties': {'frozenRowCount': 1}
                                    },
                        'fields': 'gridProperties.frozenRowCount',
                                                }
                    },
                ]
                }

        service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id, body=data).execute()

    def bold_range(self, sheet_id, start_col, end_col, start_row, end_row):
        service = self.service
        sheet = service.spreadsheets()
        spreadsheet_id = self.spreadsheet_id

        data = {"requests":
                {'repeatCell':
                 {
                'range':
                {   'sheetId': sheet_id,
                    'startColumnIndex': start_col,
                    'endColumnIndex': end_col,
                    'startRowIndex': start_row,
                    'endRowIndex': end_row
                },
                'cell':
                {'userEnteredFormat':
                    {'backgroundColor': {'red': .9,
                                         'green': .9,
                                         'blue': .9,
                                         'alpha': .1 }
                }
                },
                'fields': 'userEnteredFormat.backgroundColor.red',
                }
                }
                }

        service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id, body=data).execute()

class BookFormat(object):
    def __init__(self):

        self.service = ''
        self.spreadsheet_id = ''
        self.read_range = '' # from intake!!!!!!1
        self.shmonth = []
        self.shnames = []
        self.shyear = []

    def set_id(self, service, spreadsheet_id, read_range):
        self.service = service
        self.spreadsheet_id = spreadsheet_id
        self.read_range = read_range
        self.shmonths = ['jan', 'feb', 'mar', 'apr', 'may', 'june', 'july', 'aug', 'sep', 'oct', 'nov', 'dec']
        self.shyear = [f'{Config.current_year}']
        self.shnames = None
        self.book = 'rent sheets'

    def existing_ids(self, dict):
        # creates index of titles and ids
        idx = []
        sheet = []
        id = []
        sub = ()
        idx_dict = {}
        print("======================================")
        print("\n\n")
        for k, v in enumerate(dict):
            idx.append(k)
            sheet.append(v)
            id.append(dict[v])
            print(k, "***", v, "***", dict[v])

        print("\n")
        print("======================================")
        sub = tuple(zip(sheet, id))
        idx_list = list(zip(idx, sub))

        return idx_list

    def make_sheet_names(self):
        "Making sheet names from list . . . "

        shnames = [] # is list okay
        for i in range(len(self.shmonths)):
            shname = self.shmonths[i] + " " + self.shyear[0]
            shnames.append(shname)
        self.shnames = shnames # keeps updating internal to class
        return shnames

    def make_new_book(self, title):
        service = self.service
        data = {'properties': {'title': title}
                }
        request = service.spreadsheets().create(body=data)
        response = request.execute()
        print(response)

    def make_one_sheet(self, sheet_title):
        service = self.service
        sh_id = self.spreadsheet_id

        data = {
            'requests': [{
                'addSheet': {
                    'properties': {
                        'title': sheet_title,
                        'tabColor': {
                            'red': 0.44,
                            'green': 0.99,
                            'blue': 0.50
                        }
                    }
                }
            }]
        }

        response = service.spreadsheets().batchUpdate(
            spreadsheetId=sh_id,
            body=data
        ).execute()

    def make_sheets_from_shnames(self):
        shnames = self.shnames
        # make sheets from list
        for item in shnames:
            pprint(item)
            pprint("Taking a 30 second nap to preserve writes.  zzz....")
            sleep(30)
            pprint("Waking up . . . ")
            self.make_one_sheet(item)
        return shnames

    def del_one_sheet(self, id):
        f"Deleting sheet {id} . . . "
        service = self.service
        sh_id = self.spreadsheet_id

        data = {"requests": [
                {"deleteSheet": {"sheetId": f'{id}'}
                } ]  }
        response = service.spreadsheets().batchUpdate(
            spreadsheetId=sh_id,
            body=data
        ).execute()
        print(response)

    def make_title_list(self, shnames, titles_dict):
        print("Making clean titles list . . . ")
        titles2 = []
        print(shnames)
        for title in titles_dict.keys() :
            for shname in shnames:
                if title == shname:
                    titles2.append(shname)
        print(titles2)
        return titles2

    def clean_title_dict(self, shnames, titles_dict):
        """Validates list of sheets against titles
        prevents formatting of sheets that aren't on the list"""

        print("Cleaning titles_dict . . . ")
        t = []
        id = []
        for title in titles_dict.keys() :
            for shname in shnames:
                if title == shname:
                    t.append(shname)
                    id.append(titles_dict[title])
        clean_titles_dict = dict(zip(t, id))
        return clean_titles_dict

class DBIntake(object):

    def __init__(self):
        self.path = Config.RS_DL_FILE_PATH
        # self.dir_hrig =  Path('C:/Users/V/Google Drive/pythonwork/projects/fcv_fin/download_here')
        self.dir = None
        self.ytd_dep_book = '1vBtGLAYOIIraEiWGE-RxxnarSRSpTJr7DdsmfLXTxQs'
        self.header_names = ['Unit', "Name", "Payment_String", "Payment_Float", "Date", "Date_Code", "Trxn Id(Date_Code + Count)",]
        self.wrange_unit =  f'{Config.current_year}!A2:A200'
        self.wrange_name =  f'{Config.current_year}!B2:B200'
        self.wrange_pay =   f'{Config.current_year}!C2:C200'
        self.wrange_pay_f = f'{Config.current_year}!D2:D200'
        self.date =         f'{Config.current_year}!E2:E200'
        self.datecode =     f'{Config.current_year}!F2:F200'
        self.wrange_tid =   f'{Config.current_year}!G2:G200'
        self.wrange_id =    f'{Config.current_year}!A1:Z100'

    def load_activate_workbook(self, path):
        workbook = load_workbook(path)
        sheet = workbook.active
        pprint(f"Opening {sheet} from book {workbook} . . .  ")
        return sheet

    def make_dt(self, dict):
        pprint("Converting dates to dt and inserting in dict . . . ")
        for k, v in dict.items():
            if v['date'] != None:
                v['dt_date'] = [datetime.datetime.strptime(v['date'], "%m/%d/%Y")]
            else:
                print("NoneType not converted to datetime")

        return dict

    def containment_count(self, pay_list):
        pprint("Checking pay_list count . . .")
        total_items = 0

        for i in pay_list:
            total_items += 1

        pprint(f"Found {total_items} total_items instances in payments. Writing this to sheet . . .")

    def containment_tally(self, pay_list):
        pprint("Tallying payments for containment . . .")
        total_tally = 0
        for payment in pay_list:
            total_tally += payment

        pprint(f"Tallied {total_tally} in payments list.  Writing this to sheet . . . ")


    def intake_saink(self, sheet):
        """a modification of keetchen_saink: iterates through excel sheet and writes output to rent sheets 2020/intake in g sheets"""
        pprint("Using openpyxl to iterate through rows of excel sheet . . .")
        pprint("Generating lists of transactions et al from sheet . . . ")

        tenantname = []
        unit=[] #1
        Krent=[] #3
        actualsub=[] #7
        Trent=[] #13

        for value in sheet.iter_rows(min_row=11, min_col=1, max_col=14, values_only=True):

            if value[1] != None:
                unit.append(value[0])
                tenantname.append(value[3])
                Krent.append(value[6])
                actualsub.append(value[11])
                Trent.append(value[10])

            else:
                print(value)
                print("******")

        return tenantname, unit, Krent, actualsub, Trent

    def date_formatcol(self, service, ss_id, sheet_id, scol, endcol): # attempt to programmatically change date col format MAY NOT NEED
        sheet = service.spreadsheets()
        spreadsheet_id = ss_id

        requests = [
                {
                    "repeatCell": {
                        "range": {
                            "sheetId": sheet_id,
                            "startRowIndex": 0,
                            "endRowIndex": 100,
                            "startColumnIndex": scol,
                            "endColumnIndex": endcol
                },
                        "cell":  {
                            "userEnteredFormat": {
                                    "numberFormat": {
                                        "type": "DATE",
                                        "pattern": "mm/d/y"
                                }
                                }
                            },
                        "fields": "userEnteredFormat.numberFormat"
                        }
                        },
                ]

        body = {"requests": requests}
        response = service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id, body=body).execute()
        pprint(response)
      
       
    @staticmethod
    def DB2RS():
        # TODO
        # make a list of tenant names
        # if payee is not on list of tenant names, then we use

        service = oauth(my_scopes, 'sheet')
        format = TemplateFormatSheet()
        format.set_id(service, '1vBtGLAYOIIraEiWGE-RxxnarSRSpTJr7DdsmfLXTxQs', '2020!A2:Z100') #dbintake here
        pd.set_option('display.max_rows', None) #makes df print the entire dataframe

        """sheet select"""
        titles_dict = get_existing_sheets(service, CURRENT_YEAR_RS)
        sheet_choice, selection = show_files_as_choices(titles_dict)
        sheet_choice1 = f'{sheet_choice}!K2:K68'
        print("You've chosen to work with: " + sheet_choice1)

        """month selector"""
        ui_pick = input(f'You picked {selection} as your month choice.  If this is correct, press j, otherwise enter the 2 digit month code you would like to use. >>>')
        if ui_pick == 'j':
            month_pick = str(selection)
        else:
            month_pick = str(ui_pick)

        r = BasicBitchPayment.query.filter(BasicBitchPayment.date_code.contains(month_pick)).all()

        num_list = []
        tenant_list = [] 
        pay_list = [] 
        date_code = []
        for item in r:
            num_list.append(item.unit1)
            tenant_list.append(item.name)
            pay_list.append(item.pay_float)
            date_code.append(item.date_code)
  

        def conv(list1, type=None):
            for item in list1:
                if type == 'str':
                    yield str(item)
                else:               
                    yield float(item)
    
        # this is the contents of the database selected by date_code and arranged
        tuple_from_db = tuple(zip(tenant_list, tuple(zip(num_list, conv(pay_list), date_code))))

        # checks payment grand total: does it match sheet? 
        tally_check = 0
        new_list = []
        for count, it in enumerate(tuple_from_db, 1):
            tally_check += float(it[1][1])
            new_list.append(it)

        # pulls out non_tenant_payments: like laundry
        non_tenant_payments = 0.0
        tenant_payments_subtotal = 0.0
        for it in tuple_from_db:
            if it[0] == 'Laundry Income, Laundry Income' or it[0][0] is None:
                non_tenant_payments += float(it[1][1])
            else:
                tenant_payments_subtotal += float(it[1][1])

        df = pd.DataFrame(new_list, columns =['Name', 'Other']) 
        df[['Unit', 'Payment', 'Date Code']] = pd.DataFrame(df['Other'].tolist(), index=df.index)
        df = df.drop(labels='Other', axis=1)
        df1 = df.groupby(['Name', 'Unit'])['Payment'].sum()
    
      
        # # unit_idx is just a list of unit names as strings ie CD-A
        # unit_idx = list(conv(Unit.query.all(), type='str'))

        final_list = []
        idx_list = []
        for index, unit in enumerate(unit_idx): # indexes units from sheet
            idx_list.append(int(index))
            final_list.append(unit)

        unit_index = tuple(zip(idx_list, final_list))

        unit_index_df = pd.DataFrame(unit_index, columns= ['Rank', 'Unit'])
           
        merged_df = pd.merge(df1, unit_index_df, on='Unit', how='outer')
        df2 = merged_df.sort_values(by='Rank', axis=0)
        df2 = df2.fillna(0)
        payment_list = df2['Payment'].tolist()
     
        format.simple_batch_update(service, FileHandler().spreadsheet_testbook, sheet_choice1, payment_list, "COLUMNS") # this is the write range
    
        print(f'Total items packed from db: {count}')
        print(f'Total items amount: {tally_check}')
        print(f'Total ntp amount: {non_tenant_payments}')
        print(f'Total ntp amount: {tenant_payments_subtotal}')
    

def reconciliation_runtime():

    elif input1 == 2:
        DBIntake.DB2RS()

